import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string
import os

from localsettings import *

GREEN = 'FF00FF00'
WHITE = '00000000'
PURPLE = 'FF7777FF'

class edge_detection:

    def __init__(self):
        self.ref_wb = None
        self.ref_active_sh = None

    def reference_check(self):
        ref_path = os.getcwd() + '/' + 'Reference' + '/' + 'REF.xlsx'
        self.ref_wb = openpyxl.load_workbook(ref_path)
        self.ref_active_sh = self.ref_wb.active
        max_row = self.ref_active_sh.max_row
        max_col = self.ref_active_sh.max_column
        edge_cell_list = []
        all_adjacent_cells = set()

        for row in self.ref_active_sh.iter_rows():
            for i , cell in enumerate(row):
                if cell.fill.start_color.index == WHITE or cell.fill.start_color.index == PURPLE:
                    row_index = cell.row
                    col_index = cell.column
                    edge_cell_list = self.get_adjacent_cells(self.ref_active_sh, row_index, col_index)
                    all_adjacent_cells.update(edge_cell_list)

        return all_adjacent_cells

    def is_point_in_adjacent_cells(self, all_adjacent_cells, point):
        return point in all_adjacent_cells


    def get_adjacent_cells(self, sheet, row, column, n=5):
        indices = []
        # 위 방향으로 n개의 셀 인덱스 추가
        for i in range(n):
            if row - (i + 1) > 0:
                indices.append((row - (i + 1), column))

        # 아래 방향으로 n개의 셀 인덱스 추가
        for i in range(n):
            indices.append((row + i + 1, column))

        # 왼쪽 방향으로 n개의 셀 인덱스 추가
        for i in range(n):
            if column - (i + 1) > 0:
                indices.append((row, column - (i + 1)))

        # 오른쪽 방향으로 n개의 셀 인덱스 추가
        for i in range(n):
            indices.append((row, column + i + 1))

        return indices



    def test(self, cell):
        path = os.getcwd() + '/' + 'Reference' + '/' + 'aaa.xlsx'
        wb = openpyxl.load_workbook(path)
        active_sh = wb.active
        color = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        for aa in cell:
            active_sh.cell(aa[0], aa[1]).fill = color
        wb.save(path)

if __name__ == '__main__':
    print("aa")
    ed = edge_detection()
    all_ad_cells = ed.reference_check()
    sorted_tuple = sorted(all_ad_cells, key=lambda x: x)
    abc = ed.is_point_in_adjacent_cells(sorted_tuple, (45,15))
    ed.test(all_ad_cells)