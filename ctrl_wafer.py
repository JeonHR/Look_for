import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
import os

from localsettings import *
from ctrl_edge import *


class ctrl_wafer:
    def __init__(self):
        self.start_col_x = None
        self.start_col_y = None
        self.start_row_x = None
        self.start_row_y = None
        self.total_wafer_id_x = []
        self.total_wafer_id_y = []
        self.theothers_index = None
        self.localsettings = Local_Settings()
        self.load_settings(self.localsettings.load())
        self.edge = edge_detection()
        self.all_ad_cells = set()
        self.all_ad_cells = self.edge.reference_check()
        self.start_x = 1
        self.start_y = 2

    def load_settings(self, settings):
        self.edge_color_item = settings['EDGE_COLOR_ITEM']
        self.enable_edge = settings['ENABLE_EDGE']

    def draw_map(self, df, wafer_id, wafer_num, r_path, sbin_item, color_item, wafer_id_total_ws, wafer_id_total_wb, num_array, edge_color):
        start_li = []
        wb = wafer_id_total_wb
        ws = wb.active
        target = wb.copy_worksheet(ws)
        # sheet타이틀 설정
        name = wafer_id + '_' + str(int(float(wafer_num)))
        print('Make sheet ' + name)
        target.title = name
        sheet = wb[name]
        sep_edge = 0

        max_row = sheet.max_row
        max_col = sheet.max_column


        self.draw_legend(sheet, sbin_item, color_item)

        for drow in df.iterrows():
            x = int(float(drow[1]['Wafer Y']))
            y = int(float(drow[1]['Wafer X']))
            cell_x = max_row - (x + self.start_x)
            cell_y = y + self.start_y
            cell = sheet.cell(row=cell_x, column=cell_y)

            for num, sbin in enumerate(sbin_item):
                color = PatternFill(start_color=color_item[num], end_color=color_item[num], fill_type='solid')
                edgecolor = PatternFill(start_color=edge_color, end_color=edge_color, fill_type='solid')
                if str(drow[1]['sbin']).lower() == sbin.lower():
                    if self.edge.is_point_in_adjacent_cells(self.all_ad_cells, (cell_x, cell_y)) and self.enable_edge == True:
                        sheet.cell(row=cell_x, column=cell_y).fill = edgecolor
                        sep_edge = sep_edge + 1
                    else:
                        sheet.cell(row=cell_x, column=cell_y).fill = color
                    if cell.fill.start_color.index != 'FF7777FF':
                        if sheet.cell(row=cell_x, column=cell_y).value == None:
                            sheet.cell(row=cell_x, column=cell_y).value = 1
                        else:
                            sheet.cell(row=cell_x, column=cell_y).value = sheet.cell(row=cell_x, column=cell_y).value + 1

                    if self.edge.is_point_in_adjacent_cells(self.all_ad_cells, (cell_x, cell_y)) and self.enable_edge == True:
                        wafer_id_total_ws.cell(row=cell_x, column=cell_y).fill = edgecolor
                        self.total_edge_count = self.total_edge_count + 1
                    else:
                        wafer_id_total_ws.cell(row=cell_x, column=cell_y).fill = color
                    if cell.fill.start_color.index != 'FF7777FF':
                        if wafer_id_total_ws.cell(row=cell_x, column=cell_y).value == None:
                            wafer_id_total_ws.cell(row=cell_x, column=cell_y).value = 1
                        else:
                            wafer_id_total_ws.cell(row=cell_x, column=cell_y).value = wafer_id_total_ws.cell(row=cell_x, column=cell_y).value + 1
                    num_array[num].append(sbin)

        sheet.cell(row=max_row, column=max_col + 2).value = "Edge count : {}".format(sep_edge)
        wb.save(r_path)

    def total_draw_map(self, r_path, wafer_id, id_num):
        self.total_edge_count = 0
        start_li = []
        wafer_id_total_wb = openpyxl.load_workbook(r_path)
        wafer_id_total_ws_ref = wafer_id_total_wb["Sheet1"]
        wafer_id_total_ws = wafer_id_total_wb.copy_worksheet(wafer_id_total_ws_ref)
        # sheet타이틀 설정
        wafer_id_total_ws.title = 'Total ' + wafer_id
        wafer_id_total_ws.sheet_properties.tabColor = "00ff00"
        print('Total__' + wafer_id)

        self.max_row = wafer_id_total_ws.max_row
        self.max_col = wafer_id_total_ws.max_column

        # sheet.cell()
        for i in range(1, self.max_row + 1):
            for j in range(1, self.max_col + 1):
                if wafer_id_total_ws.cell(row=i, column=j).value == 'Wafer_ID':
                    wafer_id_total_ws.cell(row=i + 1, column=j).value = 'TOTAL'
                elif wafer_id_total_ws.cell(row=i, column=j).value == 1:
                    start_li.append(i)
                    start_li.append(j)

        return wafer_id_total_ws, wafer_id_total_wb

    def total_map_save(self, path, wafer_id_total_wb, wafer_id_total_ws, sbin_item, color_item, num_array, edge_color):
        self.draw_legend(wafer_id_total_ws, sbin_item, color_item)
        self.draw_num(wafer_id_total_ws, sbin_item, color_item, num_array)
        wafer_id_total_ws.cell(row=wafer_id_total_ws.max_row, column=wafer_id_total_ws.max_column - 4).value = "Edge count : {}".format(self.total_edge_count)
        wafer_id_total_wb.save(path)

    def draw_legend(self, ws, sbin_item, color_item):
        max_row = ws.max_row
        max_col = ws.max_column

        for i in range(len(sbin_item)):
            if sbin_item[i] == '':
                continue
            color = PatternFill(start_color=color_item[i], end_color=color_item[i],
                                fill_type='solid')
            ws.cell(i + 3, max_col + 2).fill = color
            ws.cell(i + 3, max_col + 3).value = sbin_item[i]

    def draw_num(self, ws, sbin_item, color_item, num_array):
        max_row = ws.max_row
        max_col = ws.max_column

        for i in range(len(sbin_item)):
            if sbin_item[i] == '':
                continue
            ws.cell(i + 3, max_col + 4).value = len(num_array[i]) - 1


    def check_and_save_df(self, df):
        edge_list = []
        result_path = os.getcwd() + '/' + 'Result' + '/' + 'Result_log.xlsx'
        for drow in df.iterrows():
            x = int(float(drow[1]['Wafer Y']))
            y = int(float(drow[1]['Wafer X']))
            cell_x = self.max_row - (x + self.start_x)
            cell_y = y + self.start_y
            if self.edge.is_point_in_adjacent_cells(self.all_ad_cells, (cell_x, cell_y)):
                edge = 'YES'
            else:
                edge = 'NO'
            edge_list.append(edge)
        df["edge"] = edge_list
        df.to_excel(result_path, index=False)