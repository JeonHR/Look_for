import os
import openpyxl
from openpyxl import Workbook
import sys
import time
import shutil
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill
from localsettings import *
from ctrl_wafer import *
from converter import *


class main:
    def __init__(self):
        self.localsettings = Local_Settings()
        self.load_settings(self.localsettings.load())
        self.draw = ctrl_wafer()

        self.con = converter()

        self.ref_wb = None
        self.ref_active_sh = None

        self.wafer_id = ''
        self.wafer_no = ''
        self.num_array = [[0 for col in range(1)] for row in range(len(self.sbin_item))]

    def load_settings(self, settings):
        self.test_item = settings['TEST_ITEM']
        self.sbin_item = settings['SBIN_ITEM']
        self.color_item = settings['COLOR_ITEM']
        self.edge_color_item = settings['EDGE_COLOR_ITEM']

    def read_ref(self, ref_path):
        self.ref_wb = openpyxl.load_workbook(ref_path)
        self.ref_active_sh = self.ref_wb.active

    def create_copy_excel_file(self, ref_path, result_path):
        #ref 파일 복사
        if os.path.isfile(result_path):
            os.remove(result_path)
        shutil.copyfile(ref_path, result_path)
        wb = openpyxl.load_workbook(result_path)
        wb.save(result_path)

    def make_df_ate(self):
        Log_dir = os.getcwd() + '/Con_Ex'
        filelist = self.target_csvfile(self.search(Log_dir))

        df_list = []
        #csv log 파일이 여러개 인지
        for i, file in enumerate(filelist):
            df = pd.read_csv(file, low_memory=False)
            print("read exporter_output.csv shape : {0}".format(df.shape))
            df_list.append(df)
        df_merge = pd.concat(df_list, axis=1)
        #df_merge = df_merge.set_index(df_merge.SerialNumber)     5e
        if ('upper' in df_merge.SerialNumber[0].lower()) or ('upper' in df_merge.SerialNumber[1].lower()) or \
                'upper' in df_merge.SerialNumber[2].lower() or \
                'upper' in df_merge.SerialNumber[3].lower():

            df_merge = df_merge.drop(['Upper Limit', 'Lower Limit'], axis=1)
        df_merge = df_merge.dropna(axis=1)
        return df_merge

    def check_wafer(self, df_merge, result_path):

        wafer_list = df_merge['Wafer LotID'].unique()
        print("wafer ID list -" + wafer_list)
        for k, wafer_id in  enumerate(wafer_list):
            df_wa = df_merge[df_merge['Wafer LotID'].str.contains(wafer_id)]
            df_wa = df_wa.astype({'Wafer No':'float',
                                  'Wafer X':'float',
                                  'Wafer Y':'float'})
            df_wa = df_wa.astype({'Wafer No':'str',
                                  'Wafer X':'str',
                                  'Wafer Y':'str'})
            wnum = df_wa['Wafer No'].unique()

            wafer_id_total_sheet, wafer_id_total_wb = self.draw.total_draw_map(result_path, wafer_id, k)
            for wafer_num in wnum:
                df_num = df_wa[df_wa['Wafer No'] == wafer_num]
                self.draw.draw_map(df_num, wafer_id, wafer_num, result_path, self.sbin_item, self.color_item, wafer_id_total_sheet, wafer_id_total_wb, self.num_array, self.edge_color_item)
            self.draw.total_map_save(result_path, wafer_id_total_wb, wafer_id_total_sheet, self.sbin_item, self.color_item, self.num_array, self.edge_color_item[0])
            self.num_array = [[0 for col in range(1)] for row in range(len(self.sbin_item))]

        wb = openpyxl.load_workbook(result_path)
        ws = wb['Sheet1']
        wb.remove(ws)
        wb.save(result_path)

        self.draw.check_and_save_df(df_merge)
        print("Finish---")

    def search(self, folder_name, file_list=[]):
        file_list.clear()
        for (path, dir, files) in os.walk(folder_name):
            for filename in files:
                ext = os.path.splitext(filename)[-1]
                if ext == '.csv' or ext == '.xlsx':
                    print("%s/%s" % (path, filename))
                    fullFilename = os.path.join(path, filename)
                    file_list.append(fullFilename)
        return file_list

    def target_csvfile(self, files, csvfile_list=[]):
        csvfile_list.clear()
        for file in files:
            if ".csv" in file:
                if "._" in file:
                    continue
                csvfile_list.append(file)
        return csvfile_list

    def convert_export_log(self):
        self.con.converter()

if __name__ == '__main__':
    print("Wafer Mapping automation 2.0.6 - 230809")
    m = main()
    m.convert_export_log()
    ref_path = os.getcwd() + '/' + 'Reference' + '/' + 'REF.xlsx'
    result_path = os.getcwd() + '/' + 'Result' + '/' + 'Result_draw.xlsx'
    m.read_ref(ref_path)
    m.create_copy_excel_file(ref_path, result_path)
    df = m.make_df_ate()
    m.check_wafer(df, result_path)